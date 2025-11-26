#views/asistencia
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.utils import timezone
from empleados.models import Empleado
from ubicaciones.models import Ubicaciones
from .models import Asistencia
from config.decorators import role_required
from datetime import datetime


def determinar_turno_actual():
    hora = datetime.now().hour
    if 6 <= hora < 14:
        return "06-14"
    elif 14 <= hora < 22:
        return "14-22"
    else:
        return "22-06"


# Create your views here.
@role_required(["Oficial"])
def registrar_asistencia_view(request):
    ubicaciones = Ubicaciones.objects.all().order_by('nombre')
    today = timezone.localdate().strftime('%Y-%m-%d')

    if request.method == "POST":
        id_ubicacion = request.POST.get("id_ubicacion", "").strip()
        observaciones = request.POST.get("observaciones", "").strip()

        usuario_email = request.session.get("usuario_email")
        empleado = get_object_or_404(Empleado, email=usuario_email)

        ubicacion = get_object_or_404(Ubicaciones, id_ubicacion=id_ubicacion)

        try:
            nueva_asistencia = Asistencia.objects.create(
                id_empleado=empleado,
                id_ubicacion=ubicacion,
                turno_ingreso=timezone.now(),
                observaciones=observaciones or None,
                estado='En curso'
            )
            messages.success(
                request,
                f"✅ Asistencia registrada correctamente a las {nueva_asistencia.turno_ingreso.strftime('%H:%M:%S')}.",
                extra_tags="crear"
            )
        except Exception as e:
            messages.error(request, f"⚠️ Error al registrar asistencia: {str(e)}", extra_tags="crear")

        return redirect("registrarAsistencia")

    context = {
        "ubicaciones": ubicaciones,
        "today": today
    }
    return render(request, "asistencia/registrarAsistencia.html", context)


@role_required(["Oficial"])
def asistencias_activas_view(request):
    usuario_email = request.session.get("usuario_email")
    empleado = get_object_or_404(Empleado, email=usuario_email)

    asistencias = Asistencia.objects.filter(
        id_empleado=empleado,
        estado='En curso'
    ).order_by('turno_ingreso')

    if request.method == "POST":
        accion = request.POST.get("accion")
        id_asistencia = request.POST.get("id_asistencia")
        asistencia = get_object_or_404(Asistencia, id_asistencia=id_asistencia)

        if accion == "salida":
            asistencia.turno_salida = datetime.now()
            asistencia.estado = 'Finalizado'
            asistencia.save()
            messages.success(request, f"✅ Turno finalizado a las {asistencia.turno_salida.strftime('%H:%M:%S')}", extra_tags="crear")
        elif accion == "editar":
            observaciones = request.POST.get("observaciones", "").strip()
            asistencia.observaciones = observaciones or None
            asistencia.save()
            messages.success(request, "✅ Observaciones actualizadas correctamente", extra_tags="crear")

        return redirect("consultarAsistencia")

    return render(request, "asistencia/consultarAsistencia.html", {"asistencias": asistencias})

#---------------------------------PARA EXPORTAR A EXCEL ---------------------------------------------

# --- LISTADO Y EXPORT PARA OFICIALES (pegar al final de asistencia/views.py) ---
from datetime import datetime, time
from io import BytesIO
from django.http import HttpResponse
from django.core.paginator import Paginator
from django.utils import timezone
from empleados.models import Empleado
from .models import Asistencia
from config.decorators import role_required

DATE_FORMATS = ("%Y-%m-%d", "%d/%m/%Y")

def parse_date(s: str):
    if not s:
        return None
    s = s.strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None

def make_aware_dt(d, end=False):
    if not d:
        return None
    naive = datetime.combine(d, time.max if end else time.min)
    return timezone.make_aware(naive, timezone.get_current_timezone())

def _get_empleado_from_request(request):
    email = getattr(getattr(request, "user", None), "email", None) or request.session.get("usuario_email")
    return Empleado.objects.filter(email=email).first()

def _build_filtered_qs_oficiales(request):
    empleado_actual = _get_empleado_from_request(request)

    empleado_id  = (request.GET.get("id_empleado") or "").strip()
    fecha_inicio = (request.GET.get("fecha_inicio") or "").strip()
    fecha_fin    = (request.GET.get("fecha_fin") or "").strip()

    di = parse_date(fecha_inicio)
    df = parse_date(fecha_fin)

    di_dt = make_aware_dt(di, end=False) if di else None
    df_dt = make_aware_dt(df, end=True)  if df else None

    qs = Asistencia.objects.select_related("id_empleado", "id_ubicacion").all()

   
    try:
        es_oficial = (getattr(request.user, "rol", None) == "Oficial")
    except:
        pass
    if es_oficial and empleado_actual:
        qs = qs.filter(id_empleado=empleado_actual)

    if empleado_id:
        qs = qs.filter(id_empleado__pk=empleado_id)
    if di_dt:
        qs = qs.filter(turno_ingreso__gte=di_dt)
    if df_dt:
        qs = qs.filter(turno_ingreso__lte=df_dt)

    qs = qs.order_by("-turno_ingreso")

    filtros_ctx = {
        "empleado_id": empleado_id,
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
    }
    return qs, filtros_ctx, es_oficial, empleado_actual

@role_required(["Administrador"])
def ver_asistencia_oficiales_view(request):
    qs, filtros_ctx, es_oficial, empleado_actual = _build_filtered_qs_oficiales(request)

    paginator = Paginator(qs, 25)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    if es_oficial and empleado_actual:
        empleados = Empleado.objects.filter(pk=empleado_actual.pk)
    else:
        empleados = Empleado.objects.all().order_by("nombre_completo")

    ctx = {
        "empleados": empleados,
        "asistencias": page_obj.object_list,
        "page_obj": page_obj,
        **filtros_ctx,
    }
    return render(request, "empleados/verAsistenciaOficiales.html", ctx)

@role_required(["Administrador"])
def ver_asistencia_oficiales_export(request):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment

    qs, filtros_ctx, _, _ = _build_filtered_qs_oficiales(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencias"

    headers = ["Fecha", "Empleado", "Hora Entrada", "Hora Salida", "Ubicación", "Observaciones", "Estado"]
    ws.append(headers)

    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = bold
        cell.alignment = center

    for a in qs:
        ing = timezone.localtime(a.turno_ingreso)
        sal = timezone.localtime(a.turno_salida) if a.turno_salida else None
        ws.append([
            ing.strftime("%d/%m/%Y"),
            f"{getattr(a.id_empleado, 'cedula', a.id_empleado_id)} - {getattr(a.id_empleado, 'nombre_completo', a.id_empleado_id)}",
            ing.strftime("%H:%M"),
            sal.strftime("%H:%M") if sal else "--:--",
            getattr(a.id_ubicacion, "nombre", a.id_ubicacion_id),
            a.observaciones or "",
            a.estado,
        ])

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = str(cell.value) if cell.value is not None else ""
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    nombre = ["asistencias"]
    if filtros_ctx.get("fecha_inicio"):
        nombre.append(f"ini_{filtros_ctx['fecha_inicio']}")
    if filtros_ctx.get("fecha_fin"):
        nombre.append(f"fin_{filtros_ctx['fecha_fin']}")
    if filtros_ctx.get("empleado_id"):
        nombre.append(f"emp_{filtros_ctx['empleado_id']}")
    filename = "_".join(nombre) + ".xlsx"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    resp = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp

# ------- EXPORTAR A PDF  -------
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

@role_required(["Administrador"])
def ver_asistencia_oficiales_export_pdf(request):
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    qs, filtros_ctx, _, _ = _build_filtered_qs_oficiales(request)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20
    )

    styles = getSampleStyleSheet()
    # Estilo para todas las celdas del cuerpo (texto claro)
    body_style = ParagraphStyle(
        name="BodyWhite",
        parent=styles["Normal"],
        textColor=colors.HexColor("#E5E7EB"),
        fontSize=9,
        leading=11
    )

    elems = []

    # Título
    elems.append(Paragraph("Reporte de Asistencias", styles["Title"]))

    # Línea con filtros usados
    filtros_txt = []
    if filtros_ctx.get("fecha_inicio"):
        filtros_txt.append(f"Inicio: {filtros_ctx['fecha_inicio']}")
    if filtros_ctx.get("fecha_fin"):
        filtros_txt.append(f"Fin: {filtros_ctx['fecha_fin']}")
    if filtros_ctx.get("empleado_id"):
        filtros_txt.append(f"Empleado ID: {filtros_ctx['empleado_id']}")
    if filtros_txt:
        elems.append(Paragraph(" / ".join(filtros_txt), styles["Normal"]))
    elems.append(Spacer(1, 12))

    # Datos de la tabla
    header = ["Fecha", "Empleado", "Hora Entrada", "Hora Salida", "Ubicación", "Observaciones", "Estado"]
    data = [header]

    for a in qs:
        ing = timezone.localtime(a.turno_ingreso)
        sal = timezone.localtime(a.turno_salida) if a.turno_salida else None

        empleado_str = f"{getattr(a.id_empleado, 'cedula', a.id_empleado_id)} - {getattr(a.id_empleado, 'nombre_completo', a.id_empleado_id)}"
        fecha_p = Paragraph(ing.strftime("%d/%m/%Y"), body_style)
        emp_p   = Paragraph(empleado_str, body_style)
        he_p    = Paragraph(ing.strftime("%H:%M"), body_style)
        hs_p    = Paragraph(sal.strftime("%H:%M") if sal else "--:--", body_style)
        ubi_p   = Paragraph(getattr(a.id_ubicacion, "nombre", a.id_ubicacion_id), body_style)
        obs_p   = Paragraph(a.observaciones or "", body_style)
        est_p   = Paragraph(a.estado, body_style)

        data.append([fecha_p, emp_p, he_p, hs_p, ubi_p, obs_p, est_p])

    # Anchos de columnas
    col_widths = [80, 220, 80, 80, 150, 260, 90]

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        # Encabezado
        ("BACKGROUND",      (0, 0), (-1, 0), colors.HexColor("#374151")),
        ("TEXTCOLOR",       (0, 0), (-1, 0), colors.whitesmoke),
        ("FONTNAME",        (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",        (0, 0), (-1, 0), 11),
        ("ALIGN",           (0, 0), (-1, 0), "LEFT"),
        ("VALIGN",          (0, 0), (-1, 0), "MIDDLE"),

        # Cuerpo (los Paragraph ya llevan color claro)
        ("VALIGN",          (0, 1), (-1, -1), "MIDDLE"),
        ("TOPPADDING",      (0, 1), (-1, -1), 4),
        ("BOTTOMPADDING",   (0, 1), (-1, -1), 4),
        ("LEFTPADDING",     (0, 1), (-1, -1), 6),
        ("RIGHTPADDING",    (0, 1), (-1, -1), 6),

        # Alternar fondos
        ("ROWBACKGROUNDS",  (0, 1), (-1, -1), [colors.HexColor("#0F172A"), colors.HexColor("#111827")]),

        # Rejilla
        ("GRID",            (0, 0), (-1, -1), 0.25, colors.HexColor("#4B5563")),
    ]))

    elems.append(table)
    doc.build(elems)

    pdf = buffer.getvalue()
    buffer.close()

    # Nombre de archivo con filtros
    partes = ["asistencias"]
    if filtros_ctx.get("fecha_inicio"):
        partes.append(f"ini_{filtros_ctx['fecha_inicio']}")
    if filtros_ctx.get("fecha_fin"):
        partes.append(f"fin_{filtros_ctx['fecha_fin']}")
    if filtros_ctx.get("empleado_id"):
        partes.append(f"emp_{filtros_ctx['empleado_id']}")
    filename = "_".join(partes) + ".pdf"

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
